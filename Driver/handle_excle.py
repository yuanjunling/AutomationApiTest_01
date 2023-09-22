# coding=utf-8
import openpyxl
from openpyxl import load_workbook


class HandExcel:
    def __init__(self, file_path):
        self.file_path = file_path

    def load_excel(self):
        # 加载exce
        # open_excel = openpyxl.load_workbook("E:\AutomationApiTest\Data\case_01.xlsx")
        open_excel = openpyxl.load_workbook(self.file_path)
        return open_excel

    def get_sheet_data(self, index=None):
        # 加载试验sheet的内容
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols, index=None):
        # 获取某一个单元格内容
        data = self.get_sheet_data(index).cell(row=row, column=cols).value
        return data

    def get_rows(self, index=None):
        # 获取行数
        row = self.get_sheet_data(index).max_row
        return row

    def get_rows_value(self, row, index=None):
        # 获取某一行的内容
        row_list = []
        for i in self.get_sheet_data(index)[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        """
        写入数据
        """
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save("E:\AutomationApiTest\Data\case_01.xlsx")
        # 调用此方法的时候，写入一行数据到excel最后一行

    def write_cell_content(self, array, sheet=None):
        workbook1 = load_workbook(
            "E:\AutomationApiTest\Data\Reward_User_Import_Template (1).xlsx"
        )
        if sheet == None:
            sheet = "Sheet3"
        sheet = workbook1[sheet]
        sheet.append(array)
        workbook1.save("E:\AutomationApiTest\Data\Reward_User_Import_Template (1).xlsx")
        return sheet.append(array)

    def get_columns_value(self, key=None):
        """
        获取某一列得数据
        """
        columns_list = []
        if key == None:
            key = "A"
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        """
        获取行号
        """
        num = 0
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        """
        获取excel里面所有的数据
        """
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 1))

        return data_list
